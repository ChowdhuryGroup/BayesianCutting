from skopt import Optimizer
from skopt.space import Real, Integer
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pandas
from skopt.plots import (
    plot_evaluations,
    plot_objective,
    plot_convergence,
    plot_gaussian_process,
)
import beamConstruct
import openpyxl

parameterFilepath = r"C:\Users\twardowski.6a\Documents\GlassCutting\2025-04-27\TestReadWrite.xlsx"
beamPTemplates = r"C:\Users\twardowski.6a\Documents\GlassCutting\2025-04-27\Templates"
# Define the parameter space min and max
space = [
    Real(1.5, 3.05, name="Power"),  # Pulse energy (J)
    Real(
        18, 20.1, name="FocalPosition"
    ),  # Position of slide (mm) with respect to bessel characterization
    Real(1, 150, name="scan_speed"),  # Software Scan speed (mm/s)
    Real(0.001, 0.01, name="HatchSpacing"),  # Spacing of hatch (m)
    Integer(0, 40, name="Repeats"),  # Number of times to repeat circle (unitless)
    Integer(85000, 200000, name="Compressor Setting"),  # Pharos Compressor Setting
]   


parameterSheet = pandas.read_excel(
    parameterFilepath,
    sheet_name="Parameters",
).to_numpy()


initial_parameters = []
initial_quality_factors = []

for trialNumber in range(1, len(parameterSheet)):
    initial_parameters.append(list(parameterSheet[trialNumber, 1:len(space)+1]))
    initial_quality_factors.append(parameterSheet[trialNumber, len(space)+1])

print("Input Parameters: ",initial_parameters)
print("Quality Parameters: ", initial_quality_factors)



# Define some constraints
def outputConstraints(params):
    (
        pulse_energy,
        focal_position,
        scan_speed,
        hatch_spacing,
        repeats,
        compressor_setting,
    ) = params
    # Test timelieness
    # repeats * Diameter * pi * # of hatches * # of circles
    if repeats * 0.25 * np.pi * 6 * 4 / scan_speed > 350:
        return False  # Constraint violated
    return True  # Constraint satisfied


# Initialize the optimizer, this is where you set the optimizer strategies and such
optimizer = Optimizer(
    dimensions=space,
    random_state=0,
    # acq_func="PI",
    # acq_func_kwargs={"xi": 0.1},
)  # Set random_state to None for randomization of outputs, generate next suggested parameters with random_state=0,1,2

# Feed the initial data to the optimizer
for params, quality in zip(initial_parameters, initial_quality_factors):
    if quality == 0:
        quality = -00
    optimizer.tell(params, -quality)  # Negate quality factor since we are minimizing

# Perform iterative optimization

# Get the next suggested parameter set
requested_points = 4
suggested_params = optimizer.ask(n_points=requested_points, strategy="cl_mean")
# Limit those that are outside of our time constraint
valid_suggestions = [p for p in suggested_params if outputConstraints(p)]
invalid_suggestions = [p for p in suggested_params if not outputConstraints(p)]

while len(invalid_suggestions) > 0:
    print("FEEDING INVALID SUGGESTIONS")
    for params in invalid_suggestions:
        optimizer.tell(params, 0)  # Quality factor 0 for invalid suggestions
    suggested_params = optimizer.ask(n_points=requested_points, strategy="cl_min")
    # Limit those that are outside of our time constraint
    # suggested_params = [p for p in suggested_params if outputConstraints(p)]
    valid_suggestions = [p for p in suggested_params if outputConstraints(p)]
    invalid_suggestions = [p for p in suggested_params if not outputConstraints(p)]

suggested_params = valid_suggestions

# Print the suggested parameters

tested_parameters = np.array(optimizer.Xi)
tested_quality_factors = np.array(optimizer.yi)
# print("\nAll Parameters Tested:")
# for params, quality in zip(optimizer.Xi, [-q for q in optimizer.yi]):
#    print(f"Parameters: {params}, Quality Factor: {quality:.3f}")


# Print parameters
print("\nSuggested Next Parameters:")
for i in range(len(suggested_params)):
    print("\t".join(map(str, suggested_params[i])))

#Lets write these parameters straight into the excel sheet

# Convert suggested_params to a DataFrame
new_data = pandas.DataFrame(suggested_params)
print(new_data)

# Load the workbook and select the sheet
workbook = openpyxl.load_workbook(parameterFilepath)
sheet = workbook["Parameters"]



start_trial_number = sheet.cell(row=sheet.max_row, column=1).value


# Append the new data to the sheet and create BeamConstruct Templates
for i, row in enumerate(new_data.itertuples(index=False, name=None), start=start_trial_number + 1):
    sheet.append([i] + list(row))
    beamConstruct.generateBeampFile(i,row[2],row[4],row[3],beamPTemplates)

# Save the workbook
workbook.save(parameterFilepath)

exit()
# Plot quality factor vs iteration
plt.figure(figsize=(8, 6))
plt.plot(
    range(1, len(tested_quality_factors) + 1),
    tested_quality_factors,
    marker="o",
    label="Quality Factor",
)
plt.xlabel("Iteration")
plt.ylabel("Quality Factor")
plt.title("Optimization Progress")
plt.legend()
plt.show()


# Visualizing optimization process: https://scikit-optimize.github.io/stable/auto_examples/plots/visualizing-results.html
plot_evaluations(optimizer.get_result())
plt.show()

plot_objective(optimizer.get_result())
plt.show()

plot_convergence(optimizer.get_result())
plt.show()


def plot_acquisition_2D_slice(optimizer, space, dim_x, dim_y, fixed_params):
    """
    Visualize a 2D slice of the acquisition function.

    :param optimizer: Optimizer object
    :param space: Parameter space
    :param dim_x: Index of the first varying dimension
    :param dim_y: Index of the second varying dimension
    :param fixed_params: List of fixed parameter values
    """
    # Generate grid for the two selected dimensions
    x_range = np.linspace(space[dim_x].low, space[dim_x].high, 50)
    y_range = np.linspace(space[dim_y].low, space[dim_y].high, 50)
    X, Y = np.meshgrid(x_range, y_range)

    # Prepare parameter sets for prediction
    grid_params = [fixed_params.copy() for _ in range(len(X.ravel()))]
    for i, (x, y) in enumerate(zip(X.ravel(), Y.ravel())):
        grid_params[i][dim_x] = x
        grid_params[i][dim_y] = y

    # Predict acquisition function values
    grid_params = [fixed_params.copy() for _ in range(len(X.ravel()))]
    for i, (x, y) in enumerate(zip(X.ravel(), Y.ravel())):
        grid_params[i][dim_x] = x
        grid_params[i][dim_y] = y

    # Get mean and std predictions
    mean, std = optimizer.models[-1].predict(grid_params, return_std=True)

    # Debug: Check range of mean and std
    print(f"Mean Prediction Range: Min={np.min(mean)}, Max={np.max(mean)}")
    print(f"Std Prediction Range: Min={np.min(std)}, Max={np.max(std)}")

    Z = optimizer.models[-1].predict(grid_params).reshape(X.shape)
    print(Z)
    print("Min MAx of Z")
    print(np.min(Z))
    print(np.max(Z))

    # Plot the acquisition function
    plt.figure(figsize=(8, 6))
    plt.contourf(X, Y, Z, levels=50, cmap="viridis")
    # plt.colorbar(label="Acquisition Function Value")
    plt.xlabel(space[dim_x].name)
    plt.ylabel(space[dim_y].name)
    plt.title(
        f"2D Slice of Acquisition Function: {space[dim_x].name} vs {space[dim_y].name}"
    )
    plt.show()

    plt.legend()
    # Scatterplot of tested points
    tested_x = [p[dim_x] for p in optimizer.Xi]
    tested_y = [p[dim_y] for p in optimizer.Xi]
    plt.scatter(
        tested_x,
        tested_y,
        c="red",
        marker="o",
        edgecolors="black",
        label="Tested Points",
    )
    plt.show()


def get_fixed_params(strategy="best"):
    """
    Determine fixed parameter values based on a given strategy.

    :param strategy: "best" for best-performing parameters, "average" for mean values
    :return: List of fixed parameter values
    """
    if strategy == "best":
        # Find the best parameters based on the highest quality factor
        best_index = np.argmax([-q for q in optimizer.yi])
        fixed_params = optimizer.Xi[best_index]
    elif strategy == "average":
        # Compute the average of all tested parameters
        fixed_params = np.mean(optimizer.Xi, axis=0)
    else:
        raise ValueError("Invalid strategy. Use 'best' or 'average'.")
    print(f"Fixed Parameters: {fixed_params}")

    return fixed_params


fixed_params = [0.0001, 2, 19, 75, 0.005, 10]
plot_acquisition_2D_slice(optimizer, space, 0, 3, get_fixed_params(strategy="average"))

best_index = np.argmax([-q for q in optimizer.yi])
best_params = optimizer.Xi[best_index]
best_quality = -optimizer.yi[best_index]

print("\nBest Predicted Parameters:")
print(f"  Parameters: {best_params}")
print(f"  Quality Factor: {best_quality:.3f}")

tested_params = np.array(optimizer.Xi)
df = pandas.DataFrame(tested_params, columns=[dim.name for dim in space])

plt.figure(figsize=(10, 6))
sns.histplot(df, kde=True, bins=15, alpha=0.5)
plt.xlabel("Parameter Values")
plt.ylabel("Frequency of Sampling")
plt.title("Distribution of Tested Parameter Values")
plt.show()
